"""CSV/Excel bulk candidate import — template generation, parsing,
validation. See docs/09-candidate-intake.md.
"""
import csv
import io
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook

TEMPLATE_COLUMNS = ["Full Name", "Email", "Phone", "Source", "LinkedIn URL", "Notes"]

EXAMPLE_ROWS = [
    ["Jordan Rivera", "jordan.rivera@example.com", "+1 555 0142", "LinkedIn", "linkedin.com/in/jordanrivera", ""],
    ["Priya Sharma", "priya.sharma@example.com", "+1 555 0198", "Referral", "linkedin.com/in/priyasharma", "Referred by Alex"],
]

# Case/whitespace-insensitive header → internal field key.
HEADER_MAP = {
    "full name": "full_name",
    "email": "email",
    "phone": "phone",
    "source": "source",
    "linkedin url": "linkedin_url",
    "notes": "notes",
}

MAX_ROWS = 5000

EMAIL_RE = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")


def generate_template_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(TEMPLATE_COLUMNS)
    writer.writerows(EXAMPLE_ROWS)
    return buf.getvalue().encode("utf-8-sig")  # BOM helps Excel open it with correct encoding


def generate_template_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"
    ws.append(TEMPLATE_COLUMNS)
    for row in EXAMPLE_ROWS:
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class RowLimitExceeded(Exception):
    pass


def _map_headers(raw_headers: list[str]) -> dict[int, str]:
    mapping: dict[int, str] = {}
    for i, header in enumerate(raw_headers):
        key = HEADER_MAP.get((header or "").strip().lower())
        if key:
            mapping[i] = key
    return mapping


def parse_uploaded_file(path: Path, filename: str) -> list[dict[str, str]]:
    """Returns a list of raw row dicts (field_key -> value), unmatched
    columns dropped. Raises RowLimitExceeded past MAX_ROWS."""
    ext = path.suffix.lower()
    if ext == ".csv":
        return _parse_csv(path)
    if ext == ".xlsx":
        return _parse_xlsx(path)
    raise ValueError(f"Unsupported file type: {ext}")


def _parse_csv(path: Path) -> list[dict[str, str]]:
    raw = path.read_bytes()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("windows-1252")  # common Excel-on-Windows export encoding

    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    header_map = _map_headers(rows[0])
    data_rows = rows[1:]
    if len(data_rows) > MAX_ROWS:
        raise RowLimitExceeded(f"File has {len(data_rows)} rows — the limit is {MAX_ROWS}")

    return [_row_to_dict(row, header_map) for row in data_rows if any(cell.strip() for cell in row)]


def _parse_xlsx(path: Path) -> list[dict[str, str]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return []
    header_map = _map_headers([str(c) if c is not None else "" for c in header_row])

    data_rows = list(rows_iter)
    if len(data_rows) > MAX_ROWS:
        raise RowLimitExceeded(f"File has {len(data_rows)} rows — the limit is {MAX_ROWS}")

    results = []
    for row in data_rows:
        cells = [str(c) if c is not None else "" for c in row]
        if any(cell.strip() for cell in cells):
            results.append(_row_to_dict(cells, header_map))
    return results


def _row_to_dict(cells: list[str], header_map: dict[int, str]) -> dict[str, str]:
    return {field_key: (cells[i].strip() if i < len(cells) else "") for i, field_key in header_map.items()}


def validate_row(row: dict[str, str]) -> tuple[str, list[str]]:
    """Returns (status, messages) — status is 'valid' | 'warning' | 'error'."""
    errors = []
    warnings = []

    if not row.get("full_name"):
        errors.append("Full Name is required")

    email = row.get("email")
    if email and not EMAIL_RE.match(email):
        errors.append("Email doesn't look valid")
    if not email:
        warnings.append("No email — duplicate detection will be weaker")
    if not row.get("phone"):
        warnings.append("No phone")

    if errors:
        return "error", errors
    if warnings:
        return "warning", warnings
    return "valid", []
